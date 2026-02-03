from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generate(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # Header
    ws['A1'] = data.get('company_name', '[Company Name]')
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = data.get('company_address', '')
    
    ws['E1'] = "INVOICE"
    ws['E1'].font = Font(size=20, color="0000FF")
    ws['E2'] = f"DATE: {data.get('date', '')}"
    ws['E3'] = f"INVOICE #: {data.get('invoice_no', '')}"

    # Bill To / Ship To
    ws['A5'] = "BILL TO:"
    ws['A5'].font = Font(bold=True)
    ws['A6'] = data.get('bill_to_name', '')
    ws['A7'] = data.get('bill_to_company', '')
    ws['A8'] = data.get('bill_to_address', '')

    ws['C5'] = "SHIP TO:"
    ws['C5'].font = Font(bold=True)
    ws['C6'] = data.get('ship_to_name', '')
    ws['C7'] = data.get('ship_to_company', '')
    ws['C8'] = data.get('ship_to_address', '')

    # Items Table
    headers = ['ITEM #', 'DESCRIPTION', 'QTY', 'UNIT PRICE', 'TOTAL']
    ws.append([])
    start_row = 10
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    current_row = start_row + 1
    for item in data.get('items', []):
        ws.cell(row=current_row, column=1, value=item.get('id', ''))
        ws.cell(row=current_row, column=2, value=item.get('description', ''))
        ws.cell(row=current_row, column=3, value=item.get('qty', 0))
        ws.cell(row=current_row, column=4, value=item.get('unit_price', 0.0))
        ws.cell(row=current_row, column=5, value=item.get('total', 0.0))
        current_row += 1

    # Totals
    ws.cell(row=current_row+1, column=4, value="SUBTOTAL")
    ws.cell(row=current_row+1, column=5, value=data.get('subtotal', 0.0))
    ws.cell(row=current_row+2, column=4, value="TAX RATE")
    ws.cell(row=current_row+2, column=5, value=data.get('tax_rate', 0.0))
    ws.cell(row=current_row+3, column=4, value="TOTAL")
    ws.cell(row=current_row+3, column=5, value=data.get('total_amount', 0.0))
    ws.cell(row=current_row+3, column=5).font = Font(bold=True)

    wb.save(output_path)
